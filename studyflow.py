import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, time
import re
from io import BytesIO
import json
import uuid
import random
import urllib.parse
from collections import defaultdict
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import base64

# Page config
st.set_page_config(
    page_title="FocusFlow",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Enhanced CSS with color coding and 30-minute increments
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    .stApp {
        font-family: 'Inter', sans-serif;
        background: linear-gradient(135deg, #0f0f23 0%, #1a1a2e 50%, #16213e 100%);
        min-height: 100vh;
        color: #ffffff;
    }
    
    .main-container {
        background: rgba(255, 255, 255, 0.05);
        backdrop-filter: blur(20px);
        border-radius: 20px;
        padding: 1.5rem;
        margin: 0.5rem;
        box-shadow: 0 20px 40px rgba(0, 0, 0, 0.3);
        border: 1px solid rgba(255, 255, 255, 0.1);
    }
    
    .app-title {
        text-align: center;
        font-size: 2rem;
        font-weight: 700;
        color: #00d2d3;
        margin-bottom: 1rem;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    
    .setup-card {
        background: rgba(255, 255, 255, 0.08);
        backdrop-filter: blur(15px);
        border-radius: 16px;
        padding: 1.5rem;
        margin: 1rem 0;
        border: 1px solid rgba(255, 255, 255, 0.15);
        color: #ffffff;
    }
    
    .step-number {
        display: inline-block;
        width: 35px;
        height: 35px;
        background: linear-gradient(135deg, #00d2d3, #48bb78);
        color: white;
        border-radius: 50%;
        text-align: center;
        line-height: 35px;
        font-weight: 600;
        margin-right: 12px;
        font-size: 1rem;
    }
    
    .course-card {
        background: rgba(255, 255, 255, 0.08);
        border-radius: 12px;
        padding: 1.25rem;
        margin: 0.75rem 0;
        border: 1px solid rgba(255, 255, 255, 0.15);
        color: #ffffff;
    }
    
    .course-code {
        font-size: 1.1rem;
        font-weight: 600;
        color: #00d2d3;
        margin-bottom: 0.5rem;
    }
    
    .course-name {
        font-size: 0.95rem;
        color: rgba(255, 255, 255, 0.85);
        margin-bottom: 0.75rem;
    }
    
    .class-schedule {
        background: rgba(0, 210, 211, 0.1);
        border-radius: 8px;
        padding: 0.75rem;
        margin: 0.5rem 0;
        border-left: 3px solid #00d2d3;
    }
    
    .class-schedule-item {
        color: rgba(255, 255, 255, 0.9);
        font-size: 0.9rem;
        margin: 0.25rem 0;
    }
    
    .intramural-card {
        background: rgba(230, 126, 34, 0.1);
        border: 1px solid rgba(230, 126, 34, 0.3);
        border-radius: 12px;
        padding: 1.25rem;
        margin: 1rem 0;
        color: #ffffff;
    }
    
    .intramural-card h4 {
        color: #e67e22;
        margin-bottom: 1rem;
        font-size: 1.1rem;
    }
    
    .activity-item {
        background: rgba(230, 126, 34, 0.15);
        border-radius: 8px;
        padding: 0.75rem;
        margin: 0.5rem 0;
        border-left: 3px solid #e67e22;
        color: #ffffff;
    }
    
    .wellness-info {
        background: rgba(253, 203, 110, 0.1);
        border: 1px solid rgba(253, 203, 110, 0.3);
        border-radius: 12px;
        padding: 1.25rem;
        margin: 1rem 0;
        color: #ffffff;
    }
    
    .wellness-info h4 {
        color: #fdcb6e;
        margin-bottom: 0.75rem;
        font-size: 1.1rem;
    }
    
    .wellness-info p {
        color: rgba(255, 255, 255, 0.9);
        line-height: 1.5;
        margin-bottom: 0.5rem;
    }
    
    .stats-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
        gap: 1rem;
        margin: 1.5rem 0;
    }
    
    .stat-card {
        background: rgba(255, 255, 255, 0.08);
        backdrop-filter: blur(15px);
        padding: 1.5rem;
        border-radius: 12px;
        text-align: center;
        border: 1px solid rgba(255, 255, 255, 0.15);
    }
    
    .stat-number {
        font-size: 2rem;
        font-weight: 700;
        color: #00d2d3;
        display: block;
    }
    
    .stat-label {
        font-size: 0.9rem;
        color: rgba(255, 255, 255, 0.8);
        margin-top: 0.5rem;
    }
    
    .progress-bar {
        height: 6px;
        background: rgba(255, 255, 255, 0.1);
        border-radius: 3px;
        overflow: hidden;
        margin: 1rem 0;
    }
    
    .progress-fill {
        height: 100%;
        background: linear-gradient(90deg, #00d2d3, #48bb78);
        transition: width 0.3s ease;
    }
    
    .progress-text {
        text-align: center;
        color: rgba(255, 255, 255, 0.8);
        font-size: 0.9rem;
        margin: 0.5rem 0;
    }
    
    .legend {
        display: flex;
        justify-content: center;
        gap: 1rem;
        margin: 1rem 0;
        flex-wrap: wrap;
    }
    
    .legend-item {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        font-size: 0.85rem;
        color: rgba(255, 255, 255, 0.8);
        background: rgba(255, 255, 255, 0.05);
        padding: 0.5rem 0.75rem;
        border-radius: 20px;
    }
    
    .legend-color {
        width: 16px;
        height: 10px;
        border-radius: 2px;
    }
    
    .week-nav {
        display: flex;
        justify-content: center;
        align-items: center;
        gap: 1rem;
        margin: 1.5rem 0;
    }
    
    .week-title {
        font-size: 1.1rem;
        font-weight: 600;
        color: #00d2d3;
        min-width: 150px;
        text-align: center;
    }
    
    /* Button styling - Clean black text on light backgrounds */
    .stButton > button {
        border: none !important;
        border-radius: 50px !important;
        padding: 0.75rem 1.5rem !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        font-family: 'Inter', sans-serif !important;
        transition: all 0.3s ease !important;
        width: 100% !important;
        background: rgba(255, 255, 255, 0.9) !important;
        color: #2d3748 !important;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1) !important;
        border: 2px solid rgba(0, 210, 211, 0.3) !important;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 25px rgba(0, 0, 0, 0.15) !important;
        background: rgba(255, 255, 255, 1) !important;
        color: #2d3748 !important;
        border: 2px solid #00d2d3 !important;
    }
    
    .stDownloadButton > button {
        border: none !important;
        border-radius: 50px !important;
        padding: 0.75rem 1.5rem !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        font-family: 'Inter', sans-serif !important;
        transition: all 0.3s ease !important;
        width: 100% !important;
        background: rgba(255, 255, 255, 0.9) !important;
        color: #2d3748 !important;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1) !important;
        border: 2px solid rgba(0, 184, 148, 0.3) !important;
    }
    
    .stDownloadButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 25px rgba(0, 0, 0, 0.15) !important;
        background: rgba(255, 255, 255, 1) !important;
        color: #2d3748 !important;
        border: 2px solid #00b894 !important;
    }
    
    /* Fix button text visibility */
    .stButton button p, .stDownloadButton button p {
        color: #2d3748 !important;
        font-weight: 600 !important;
    }
    
    /* Form styling with labels back to white */
    .stSelectbox label, .stTextInput label, .stSlider label, .stTimeInput label {
        color: #ffffff !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        font-family: 'Inter', sans-serif !important;
        margin-bottom: 0.5rem !important;
    }
    
    .stSelectbox > div > div {
        background: rgba(255, 255, 255, 0.95) !important;
        border: 2px solid rgba(200, 200, 200, 0.5) !important;
        border-radius: 12px !important;
        color: #2d3748 !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        padding: 0.75rem 1rem !important;
    }
    
    .stSelectbox > div > div > div {
        color: #2d3748 !important;
        font-weight: 600 !important;
    }
    
    /* Fix dropdown options and selected text */
    .stSelectbox div[data-baseweb="select"] > div {
        color: #2d3748 !important;
        background: rgba(255, 255, 255, 0.95) !important;
        font-weight: 600 !important;
    }
    
    .stSelectbox [data-baseweb="select"] {
        background-color: rgba(255, 255, 255, 0.95) !important;
        color: #2d3748 !important;
    }
    
    /* Dropdown menu options */
    .stSelectbox [role="option"] {
        color: #2d3748 !important;
        background-color: #ffffff !important;
        font-weight: 500 !important;
    }
    
    .stSelectbox [role="option"]:hover {
        background-color: #f7fafc !important;
        color: #2d3748 !important;
    }
    
    /* Selected option in dropdown */
    .stSelectbox [aria-selected="true"] {
        background-color: #e2e8f0 !important;
        color: #2d3748 !important;
        font-weight: 600 !important;
    }
    
    /* Main selectbox text (what you see when closed) */
    .stSelectbox span {
        color: #2d3748 !important;
        font-weight: 600 !important;
    }
    
    /* Ensure all selectbox text is visible */
    .stSelectbox * {
        color: #2d3748 !important;
    }
    
    .stTextInput > div > div > input {
        background: rgba(255, 255, 255, 0.95) !important;
        border: 2px solid rgba(255, 255, 255, 0.3) !important;
        border-radius: 12px !important;
        color: #2d3748 !important;
        font-weight: 500 !important;
        font-family: 'Inter', sans-serif !important;
        padding: 0.75rem 1rem !important;
    }
    
    .stFileUploader > div > div {
        background: rgba(255, 255, 255, 0.15) !important;
        border: 2px dashed rgba(255, 255, 255, 0.4) !important;
        border-radius: 20px !important;
        color: #ffffff !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        padding: 1.5rem !important;
        text-align: center !important;
    }
    
    .stFileUploader label {
        color: #ffffff !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        font-family: 'Inter', sans-serif !important;
    }
    
    @media (max-width: 768px) {
        .app-title {
            font-size: 1.6rem;
        }
        
        .main-container {
            margin: 0.25rem;
            padding: 1rem;
        }
        
        .legend {
            flex-direction: column;
            gap: 0.75rem;
        }
        
        .week-nav {
            flex-direction: column;
            gap: 0.75rem;
        }
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'step' not in st.session_state:
    st.session_state.step = 1
if 'courses' not in st.session_state:
    st.session_state.courses = []
if 'intramurals' not in st.session_state:
    st.session_state.intramurals = []
if 'current_week' not in st.session_state:
    st.session_state.current_week = 0
if 'user_data' not in st.session_state:
    st.session_state.user_data = {}
if 'schedule_ready' not in st.session_state:
    st.session_state.schedule_ready = False
if 'final_schedule' not in st.session_state:
    st.session_state.final_schedule = None

def parse_time_string(time_str):
    """Robust time parser that handles virtually any reasonable time format"""
    if not time_str or str(time_str).lower() == 'n/a':
        return None
    
    # Remove days from the string if present
    time_part = time_str
    if any(day in time_str.upper() for day in ['M', 'T', 'W', 'R', 'F', 'S']):
        parts = time_str.split(' ')
        if len(parts) > 1:
            time_part = ' '.join(parts[1:])  # Take everything after the days
    
    # Normalize the string: clean spaces, make lowercase
    time_part = re.sub(r'\s+', ' ', time_part.strip().lower())
    
    # Try to find any time range pattern with maximum flexibility
    # This mega-regex handles almost any reasonable time format:
    # - With/without minutes (10am, 10:30am)
    # - With/without spaces around dash (10am-12pm, 10am - 12pm, 10am- 12pm)
    # - Mixed am/pm (10am-2pm, 10:30am-11:45pm)
    # - With/without am/pm on first time (10-11am, 10am-11am)
    # - Various separators (-, to, until)
    
    time_pattern = r'''
        (\d{1,2})                    # Start hour (1-12)
        (?::(\d{2}))?                # Optional start minutes (:30)
        \s*                          # Optional space
        (am|pm)?                     # Optional AM/PM for start time
        \s*                          # Optional space
        (?:-|to|until|\s+to\s+)      # Separator (dash, "to", "until")
        \s*                          # Optional space
        (\d{1,2})                    # End hour (1-12)
        (?::(\d{2}))?                # Optional end minutes (:30)
        \s*                          # Optional space
        (am|pm)                      # Required AM/PM for end time (or we'll infer)
    '''
    
    match = re.search(time_pattern, time_part, re.VERBOSE)
    
    if match:
        start_hour = int(match.group(1))
        start_min = match.group(2) or "00"
        start_period = match.group(3)
        end_hour = int(match.group(4))
        end_min = match.group(5) or "00"
        end_period = match.group(6)
        
        # Smart period inference
        if not start_period:
            # If start has no period, infer from context
            if start_hour < end_hour and end_period.lower() == 'am':
                start_period = 'am'
            elif start_hour > end_hour and end_period.lower() == 'pm':
                start_period = 'pm'  # e.g., "11pm-1am" (next day)
            elif start_hour <= 12 and end_hour <= 12:
                start_period = end_period  # Same period
            else:
                start_period = 'am' if start_hour < 12 else 'pm'
        
        # Format the times properly
        start_time = f"{start_hour}:{start_min} {start_period.upper()}"
        end_time = f"{end_hour}:{end_min} {end_period.upper()}"
        
        return start_time, end_time
    
    # Fallback: try a simpler pattern for edge cases
    simple_pattern = r'(\d{1,2}):?(\d{2})?\s*(am|pm)?\s*[-to]\s*(\d{1,2}):?(\d{2})?\s*(am|pm)?'
    simple_match = re.search(simple_pattern, time_part)
    
    if simple_match:
        start_hour = simple_match.group(1)
        start_min = simple_match.group(2) or "00"
        start_period = simple_match.group(3) or "AM"
        end_hour = simple_match.group(4)
        end_min = simple_match.group(5) or "00"
        end_period = simple_match.group(6) or start_period
        
        return f"{start_hour}:{start_min} {start_period.upper()}", f"{end_hour}:{end_min} {end_period.upper()}"
    
    return None

def parse_days_string(schedule_str):
    """Parse days from string like 'M,W,F 10:40-11:30am', 'TR 2:00-3:30pm', or 'T,R 2:00-3:30pm'"""
    if not schedule_str or str(schedule_str).lower() == 'n/a':
        return []
    
    days_map = {
        'M': 'Monday',
        'T': 'Tuesday', 
        'W': 'Wednesday',
        'R': 'Thursday',  # R = Thursday
        'F': 'Friday',
        'S': 'Saturday',
        'U': 'Sunday'
    }
    
    days = []
    # Get just the days part (before the time)
    days_part = schedule_str.split()[0] if ' ' in schedule_str else schedule_str
    
    # Handle both "TR" and "T,R" formats
    if ',' in days_part:
        # Format like "T,R" or "M,W,F"
        day_letters = days_part.replace(',', '').replace(' ', '')
    else:
        # Format like "TR" or "MWF"
        day_letters = days_part.replace(' ', '')
    
    # Process each character as a day abbreviation
    for letter in day_letters.upper():
        if letter in days_map:
            full_day = days_map[letter]
            if full_day not in days:  # Avoid duplicates
                days.append(full_day)
    
    return days

def create_excel_template():
    """Create a sample Excel template for course data"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sample courses
    sample_courses = [
        {
            'name': 'Course 1',
            'data': [
                ['Course title', 'Biology 101 - Introduction to Biology'],
                ['Course Lecture Schedule', 'M,W,F 9:00-10:00am'],
                ['Lecture location', 'Science Building Room 101'],
                ['When in lab', 'T 2:00-4:00pm'],
                ['Where is lab', 'Biology Lab Room 205'],
                ['When is recitation', 'R 10:00-11:00am'],
                ['Where is recitation', 'Study Hall Room 15'],
                ['Suggested daily study time', '45 minutes'],
                ['', ''],
                ['Assignments:', ''],
                ['Assignment 1', 'Lab Report 1', '9/15/2024', 'Lab Report'],
                ['Assignment 2', 'Midterm Exam', '10/5/2024', 'Exam'],
                ['Assignment 3', 'Research Paper', '11/20/2024', 'Paper']
            ]
        },
        {
            'name': 'Course 2', 
            'data': [
                ['Course title', 'Chemistry 101 - General Chemistry'],
                ['Course Lecture Schedule', 'M,W,F 10:30-11:30am'],
                ['Lecture location', 'Chemistry Building Room 150'],
                ['When in lab', 'R 1:00-3:00pm'],
                ['Where is lab', 'Chemistry Lab Room 220'],
                ['When is recitation', 'N/A'],
                ['Where is recitation', 'N/A'],
                ['Suggested daily study time', '60 minutes'],
                ['', ''],
                ['Assignments:', ''],
                ['Assignment 1', 'Problem Set 1', '9/10/2024', 'Homework'],
                ['Assignment 2', 'Lab Practical', '10/12/2024', 'Lab'],
                ['Assignment 3', 'Final Exam', '12/10/2024', 'Exam']
            ]
        },
        {
            'name': 'Course 3',
            'data': [
                ['Course title', 'Mathematics 201 - Calculus I'],
                ['Course Lecture Schedule', 'T,R 11:00-12:30pm'],
                ['Lecture location', 'Math Building Room 301'],
                ['When in lab', 'N/A'],
                ['Where is lab', 'N/A'],
                ['When is recitation', 'F 2:00-3:00pm'],
                ['Where is recitation', 'Math Building Room 105'],
                ['Suggested daily study time', '90 minutes'],
                ['', ''],
                ['Assignments:', ''],
                ['Assignment 1', 'Weekly Quiz 1', '9/8/2024', 'Quiz'],
                ['Assignment 2', 'Midterm 1', '10/1/2024', 'Exam'],
                ['Assignment 3', 'Project', '11/15/2024', 'Project']
            ]
        }
    ]
    
    # Create sheets with styling
    for course_info in sample_courses:
        ws = wb.create_sheet(title=course_info['name'])
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Add data
        for row_idx, row_data in enumerate(course_info['data'], 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                
                # Style field names
                if col_idx == 1 and value and 'Course' in str(value):
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def load_config_from_json(json_data):
    """Load configuration from JSON data"""
    try:
        config = json.loads(json_data)
        
        # Load courses
        if 'courses' in config:
            st.session_state.courses = config['courses']
        
        # Load intramurals  
        if 'intramurals' in config:
            st.session_state.intramurals = config['intramurals']
        
        # Load user data
        if 'user_data' in config:
            st.session_state.user_data = config['user_data']
        
        # Mark as having processed file data
        st.session_state.file_processed = True
        
        return True
    except Exception as e:
        return False

def parse_excel_course_file(file):
    """Parse the Excel file with the specified format using pandas - memory optimized"""
    try:
        # Read only the first few sheets to reduce memory usage
        excel_data = pd.read_excel(file, sheet_name=None, header=None, nrows=50)
        courses = []
        
        # Process each sheet (each course) - limit to first 5 course sheets
        course_count = 0
        for sheet_name, df in excel_data.items():
            # Only process sheets that start with "Course" and limit to 5 courses max
            if sheet_name.startswith('Course') and course_count < 5:
                course_data = {}
                assignments = []
                
                # Safety check - ensure dataframe has enough rows and columns
                if df.shape[0] < 3 or df.shape[1] < 2:
                    continue
                
                # Check if the sheet actually has course data (check only first 5 rows)
                has_course_data = False
                for index in range(min(5, df.shape[0])):
                    try:
                        if index < df.shape[0] and df.shape[1] >= 2:
                            field_val = df.iloc[index, 0]
                            if pd.notna(field_val) and 'Course title' in str(field_val):
                                has_course_data = True
                                break
                    except:
                        continue
                
                if not has_course_data:
                    continue
                
                # Read course information (only first 20 rows)
                max_rows = min(20, df.shape[0])
                
                for index in range(max_rows):
                    try:
                        if index < df.shape[0] and df.shape[1] >= 2:
                            field_val = df.iloc[index, 0]
                            value_val = df.iloc[index, 1]
                            
                            if pd.notna(field_val) and pd.notna(value_val):
                                field = str(field_val).strip()
                                value = str(value_val).strip()
                                
                                if 'Course title' in field:
                                    course_data['name'] = value
                                elif 'Course Lecture Schedule' in field:
                                    course_data['lecture_schedule'] = value
                                elif 'Lecture location' in field:
                                    course_data['lecture_location'] = value
                                elif 'When in lab' in field:
                                    course_data['lab_schedule'] = value
                                elif 'Where is lab' in field:
                                    course_data['lab_location'] = value
                                elif 'When is recitation' in field:
                                    course_data['recitation_schedule'] = value
                                elif 'Where is recitation' in field:
                                    course_data['recitation_location'] = value
                                elif 'Suggested daily study time' in field:
                                    course_data['daily_study_time'] = value
                    except:
                        continue
                
                # Only continue if we found a course name
                if 'name' not in course_data:
                    continue
                
                # Generate course code from name
                if 'name' in course_data:
                    # Improved code generation - use first letters, not last
                    name_upper = course_data['name'].upper()
                    
                    # Try to extract standard course format first (e.g., "BIO 3455", "CHEM 101")
                    code_match = re.search(r'([A-Z]{2,4})\s*(\d{3,4})', name_upper)
                    if code_match:
                        course_data['code'] = f"{code_match.group(1)}{code_match.group(2)}"
                    else:
                        # Fallback: use first letters of significant words
                        # Remove common words and split into words
                        words = re.findall(r'[A-Z][A-Z]*', name_upper.replace('INTRODUCTION TO', '').replace('GENERAL', '').replace('BASIC', ''))
                        
                        if words:
                            # Take first 3-4 letters from the first significant word
                            first_word = words[0]
                            if len(first_word) >= 4:
                                code_prefix = first_word[:4]  # First 4 letters
                            else:
                                code_prefix = first_word[:3]  # First 3 letters if word is short
                            
                            # Look for numbers in the name
                            number_match = re.search(r'(\d{3,4})', name_upper)
                            if number_match:
                                course_data['code'] = f"{code_prefix}{number_match.group(1)}"
                            else:
                                course_data['code'] = code_prefix
                        else:
                            # Ultimate fallback - first 4 characters of cleaned name
                            cleaned_name = re.sub(r'[^A-Z0-9]', '', name_upper)
                            course_data['code'] = cleaned_name[:4] if len(cleaned_name) >= 4 else cleaned_name
                
                # Parse class schedules - handle all types (lecture, lab, recitation)
                class_schedule = []
                
                # Lecture schedule
                if course_data.get('lecture_schedule', '').lower() != 'n/a':
                    schedule_str = course_data.get('lecture_schedule', '')
                    if schedule_str:
                        days = parse_days_string(schedule_str)
                        times = parse_time_string(schedule_str)
                        if days and times:
                            class_schedule.append({
                                'days': days,
                                'start_time': times[0],
                                'end_time': times[1],
                                'type': 'Lecture',
                                'location': course_data.get('lecture_location', '')
                            })
                
                # Lab schedule
                if course_data.get('lab_schedule', '').lower() != 'n/a' and course_data.get('lab_schedule', ''):
                    schedule_str = course_data.get('lab_schedule', '')
                    if schedule_str:
                        days = parse_days_string(schedule_str)
                        times = parse_time_string(schedule_str)
                        if days and times:
                            class_schedule.append({
                                'days': days,
                                'start_time': times[0],
                                'end_time': times[1],
                                'type': 'Lab',
                                'location': course_data.get('lab_location', '')
                            })
                
                # Recitation schedule
                if course_data.get('recitation_schedule', '').lower() != 'n/a' and course_data.get('recitation_schedule', ''):
                    schedule_str = course_data.get('recitation_schedule', '')
                    if schedule_str:
                        days = parse_days_string(schedule_str)
                        times = parse_time_string(schedule_str)
                        if days and times:
                            class_schedule.append({
                                'days': days,
                                'start_time': times[0],
                                'end_time': times[1],
                                'type': 'Recitation',
                                'location': course_data.get('recitation_location', '')
                            })
                
                course_data['class_schedule'] = class_schedule
                
                # Simplified: Don't process assignments from Excel
                # Focus on class schedules and time management only
                course_data['assignments'] = []
                
                # Only add course if we got some basic info
                if 'name' in course_data:
                    courses.append(course_data)
                    course_count += 1
                
                # Clear the dataframe to free memory
                del df
        
        # Clear excel_data to free memory
        del excel_data
        
        return courses
    
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")
        return []

def generate_time_slots():
    """Generate 30-minute time slots"""
    time_slots = []
    for hour in range(6, 24):  # 6 AM to 11:30 PM
        for minute in [0, 30]:
            time_obj = time(hour, minute)
            if hour == 0:
                time_str = f"12:{minute:02d} AM"
            elif hour < 12:
                time_str = f"{hour}:{minute:02d} AM"
            elif hour == 12:
                time_str = f"12:{minute:02d} PM"
            else:
                time_str = f"{hour-12}:{minute:02d} PM"
            time_slots.append(time_str)
    return time_slots

def time_to_slot_index(time_str):
    """Convert time string to slot index with better parsing"""
    time_slots = generate_time_slots()
    
    # First try exact match
    try:
        return time_slots.index(time_str)
    except ValueError:
        pass
    
    # Try to parse and convert the time
    try:
        # Remove extra spaces
        time_str = time_str.strip()
        
        # Handle formats like "5:00 PM", "17:00", etc.
        if 'PM' in time_str.upper() or 'AM' in time_str.upper():
            # 12-hour format
            time_part = time_str.replace('PM', '').replace('AM', '').replace('pm', '').replace('am', '').strip()
            is_pm = 'PM' in time_str.upper() or 'pm' in time_str
            
            if ':' in time_part:
                hour, minute = map(int, time_part.split(':'))
            else:
                hour = int(time_part)
                minute = 0
            
            if is_pm and hour != 12:
                hour += 12
            elif not is_pm and hour == 12:
                hour = 0
        else:
            # 24-hour format
            if ':' in time_str:
                hour, minute = map(int, time_str.split(':'))
            else:
                hour = int(time_str)
                minute = 0
        
        # Convert to standard format and find in slots
        if hour == 0:
            standard_time = f"12:{minute:02d} AM"
        elif hour < 12:
            standard_time = f"{hour}:{minute:02d} AM"
        elif hour == 12:
            standard_time = f"12:{minute:02d} PM"
        else:
            standard_time = f"{hour-12}:{minute:02d} PM"
        
        # Try to find this time in slots
        try:
            return time_slots.index(standard_time)
        except ValueError:
            # Find closest time slot
            target_minutes = hour * 60 + minute
            closest_index = 0
            closest_diff = float('inf')
            
            for i, slot in enumerate(time_slots):
                slot_hour = int(slot.split(':')[0])
                slot_min = int(slot.split(':')[1].split()[0])
                if 'PM' in slot and slot_hour != 12:
                    slot_hour += 12
                elif 'AM' in slot and slot_hour == 12:
                    slot_hour = 0
                
                slot_minutes = slot_hour * 60 + slot_min
                diff = abs(slot_minutes - target_minutes)
                
                if diff < closest_diff:
                    closest_diff = diff
                    closest_index = i
            
            return closest_index
    
    except Exception as e:
        print(f"Error parsing time {time_str}: {e}")
        return 0

def generate_weekly_schedule(courses, intramurals, preferences):
    """Generate a structured weekly schedule template with 30-minute increments"""
    time_slots = generate_time_slots()
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    # Use the selected start date from preferences
    start_date = preferences.get('start_date', datetime.now().date())
    schedule_type = preferences.get('schedule_type', '📅 Current Week (this week)')
    
    # For template schedule, use generic dates
    if schedule_type == "📊 Template Week (reusable)":
        # Use a generic Monday as start date for template
        today = datetime.now().date()
        days_since_monday = today.weekday()
        start_date = today - timedelta(days=days_since_monday)
    
    # Generate just one week template
    weekly_schedule = {}
    
    for day_index, day in enumerate(days):
        daily_schedule = {}
        is_weekend = day in ["Saturday", "Sunday"]
        current_day_date = start_date + timedelta(days=day_index)
        
        # Initialize all slots as free
        for time_slot in time_slots:
            daily_schedule[time_slot] = {"activity": "Free Time", "type": "free", "date": current_day_date}
        
        # STEP 1: Add class times FIRST - HIGHEST PRIORITY
        for course in courses:
            if 'class_schedule' in course:
                for class_time in course['class_schedule']:
                    if day in class_time['days']:
                        start_time = class_time['start_time']
                        end_time = class_time['end_time']
                        class_type = class_time['type']
                        location = class_time.get('location', '')
                        
                        # Debug: Print class time info
                        print(f"Processing {course['code']} {class_type} on {day}: {start_time} - {end_time}")
                        
                        # Convert times and fill schedule - IMPROVED PARSING
                        try:
                            # More robust time parsing
                            def parse_time_better(time_str):
                                # Remove extra spaces and normalize
                                time_str = time_str.strip()
                                
                                # Handle formats like "1:55PM", "1:55 PM", "13:55"
                                if 'PM' in time_str.upper() or 'AM' in time_str.upper():
                                    # 12-hour format
                                    time_part = time_str.replace('PM', '').replace('AM', '').replace('pm', '').replace('am', '').strip()
                                    is_pm = 'PM' in time_str.upper() or 'pm' in time_str
                                    
                                    if ':' in time_part:
                                        hour, minute = map(int, time_part.split(':'))
                                    else:
                                        hour = int(time_part)
                                        minute = 0
                                    
                                    if is_pm and hour != 12:
                                        hour += 12
                                    elif not is_pm and hour == 12:
                                        hour = 0
                                else:
                                    # 24-hour format or just hour
                                    if ':' in time_str:
                                        hour, minute = map(int, time_str.split(':'))
                                    else:
                                        hour = int(time_str)
                                        minute = 0
                                
                                return hour, minute
                            
                            start_hour, start_min = parse_time_better(start_time)
                            end_hour, end_min = parse_time_better(end_time)
                            
                            print(f"Parsed times: {start_hour}:{start_min:02d} - {end_hour}:{end_min:02d}")
                            
                            # Round down start time to nearest 30 minutes (start earlier)
                            if start_min > 0 and start_min <= 30:
                                start_min = 0  # Round down to :00
                            elif start_min > 30:
                                start_min = 30  # Round down to :30
                            
                            # Round up end time to nearest 30 minutes (end later)
                            if end_min > 0 and end_min <= 30:
                                end_min = 30  # Round up to :30
                            elif end_min > 30:
                                end_min = 0
                                end_hour += 1  # Round up to next hour
                            
                            print(f"Rounded times: {start_hour}:{start_min:02d} - {end_hour}:{end_min:02d}")
                            
                            # Find corresponding time slots and fill them
                            for time_slot in time_slots:
                                slot_hour = int(time_slot.split(':')[0])
                                slot_min = int(time_slot.split(':')[1].split()[0])
                                if 'PM' in time_slot and slot_hour != 12:
                                    slot_hour += 12
                                elif 'AM' in time_slot and slot_hour == 12:
                                    slot_hour = 0
                                
                                slot_time = slot_hour * 60 + slot_min
                                start_time_min = start_hour * 60 + start_min
                                end_time_min = end_hour * 60 + end_min
                                
                                if start_time_min <= slot_time < end_time_min:
                                    activity_name = f"{course['code']} - {class_type}"
                                    if location:
                                        activity_name += f" ({location})"
                                    daily_schedule[time_slot] = {
                                        "activity": activity_name,
                                        "type": "class",
                                        "date": current_day_date
                                    }
                                    print(f"Added class to slot: {time_slot}")
                        except Exception as e:
                            print(f"Error parsing time for {course['code']} {class_type}: {e}")
                            continue
        
        # STEP 2: Add intramural activities SECOND (student-specified times get priority)
        for intramural in intramurals:
            if intramural.get('scheduled') and day in intramural.get('days', []):
                start_time = intramural.get('start_time', '5:00 PM')
                duration = intramural.get('duration', 90)  # minutes
                
                print(f"Processing intramural: {intramural['name']} on {day} at {start_time} for {duration} minutes")
                
                try:
                    start_slot = time_to_slot_index(start_time)
                    slots_needed = max(1, duration // 30)  # Ensure at least 1 slot
                    
                    print(f"Start slot index: {start_slot}, Slots needed: {slots_needed}")
                    
                    for i in range(start_slot, min(start_slot + slots_needed, len(time_slots))):
                        if i < len(time_slots):
                            slot_time = time_slots[i]
                            print(f"Trying to add activity to slot: {slot_time} (current type: {daily_schedule[slot_time]['type']})")
                            
                            # Add intramural regardless of what's there (except classes)
                            if daily_schedule[slot_time]["type"] != "class":
                                daily_schedule[slot_time] = {
                                    "activity": f"{intramural['name']} - {intramural['type']}",
                                    "type": "activity",
                                    "date": current_day_date
                                }
                                print(f"Successfully added activity to slot: {slot_time}")
                            else:
                                print(f"Slot {slot_time} blocked by class: {daily_schedule[slot_time]['activity']}")
                except Exception as e:
                    print(f"Error processing intramural {intramural['name']}: {e}")
                    continue
        
        # STEP 3: Add sleep schedule and end-of-day "Go to Sleep" (only in free slots)
        wake_time = preferences.get('wake_time', 8)
        bedtime = preferences.get('bedtime', 23)  # Using 24-hour format now
        
        # Fill sleep times - continuous sleep from bedtime to wake time
        for time_slot in time_slots:
            if daily_schedule[time_slot]["type"] == "free":  # Only if free
                hour = int(time_slot.split(':')[0])
                is_pm = 'PM' in time_slot
                is_am = 'AM' in time_slot
                
                if is_pm and hour != 12:
                    hour += 12
                elif is_am and hour == 12:
                    hour = 0
                
                # Determine if this hour should be sleep
                should_be_sleep = False
                
                if bedtime >= 22:  # Bedtime is 10 PM or later (22, 23)
                    # Sleep from bedtime until wake_time next day
                    if hour >= bedtime or hour < wake_time:
                        should_be_sleep = True
                elif bedtime <= 2:  # Bedtime is early morning (1 AM, 2 AM)
                    bedtime_24 = bedtime + 24  # Convert to 24+ for comparison
                    if hour >= 22 or hour <= bedtime or hour < wake_time:
                        should_be_sleep = True
                
                if should_be_sleep:
                    daily_schedule[time_slot] = {"activity": "Sleep", "type": "sleep", "date": current_day_date}
        
        # Add "Go to Sleep" slot 30 minutes before bedtime (only if it's currently free)
        go_to_sleep_time = None
        if bedtime >= 22:  # If bedtime is 10 PM (22) or later
            if bedtime == 22:  # 10 PM
                go_to_sleep_time = "9:30 PM"
            elif bedtime == 23:  # 11 PM
                go_to_sleep_time = "10:30 PM"
        elif bedtime == 1:  # 1 AM (stored as 25)
            go_to_sleep_time = "12:30 AM"
        elif bedtime == 2:  # 2 AM (stored as 26)
            go_to_sleep_time = "1:30 AM"
        
        if go_to_sleep_time and go_to_sleep_time in daily_schedule and daily_schedule[go_to_sleep_time]["type"] == "free":
            daily_schedule[go_to_sleep_time] = {"activity": "Go to Sleep", "type": "sleep_prep", "date": current_day_date}
        
        # STEP 4: Add meals (only in free slots, but try to be flexible)
        meal_times = {
            "8:00 AM": "Breakfast",
            "12:00 PM": "Lunch", 
            "6:00 PM": "Dinner"
        }
        
        # Try primary meal times first
        for meal_time, meal_name in meal_times.items():
            if meal_time in daily_schedule and daily_schedule[meal_time]["type"] == "free":
                daily_schedule[meal_time] = {"activity": meal_name, "type": "meal", "date": current_day_date}
        
        # If dinner was blocked by activities, try to find alternative dinner time
        dinner_scheduled = any(slot["activity"] == "Dinner" for slot in daily_schedule.values())
        if not dinner_scheduled:
            # Try alternative dinner times
            alt_dinner_times = ["5:30 PM", "6:30 PM", "7:00 PM", "7:30 PM"]
            for alt_time in alt_dinner_times:
                if alt_time in daily_schedule and daily_schedule[alt_time]["type"] == "free":
                    daily_schedule[alt_time] = {"activity": "Dinner", "type": "meal", "date": current_day_date}
                    break
        
        # STEP 5: Add study sessions based on course requirements (only in free slots)
        course_study_requirements = {}
        total_daily_study_needed = 0
        
        # Parse study time requirements for each course
        for course in courses:
            study_time_str = course.get('daily_study_time', '30 min')
            # Extract minutes from strings like "2 hours", "90 minutes", "1.5 hours", "2"
            if 'hour' in study_time_str.lower():
                hours_match = re.search(r'(\d+\.?\d*)', study_time_str)
                if hours_match:
                    hours = float(hours_match.group(1))
                    minutes = int(hours * 60)
                else:
                    minutes = 30  # default
            elif 'min' in study_time_str.lower():
                min_match = re.search(r'(\d+)', study_time_str)
                minutes = int(min_match.group(1)) if min_match else 30
            else:
                # Just a number - assume it's hours if >= 2, minutes if < 2
                num_match = re.search(r'(\d+\.?\d*)', study_time_str)
                if num_match:
                    num = float(num_match.group(1))
                    if num >= 2:
                        minutes = int(num * 60)  # assume hours
                    else:
                        minutes = int(num * 60) if num < 2 else int(num)  # assume hours or minutes
                else:
                    minutes = 30  # default
            
            # Convert to 30-minute sessions needed
            sessions_needed = max(1, round(minutes / 30))
            course_study_requirements[course['code']] = sessions_needed
            total_daily_study_needed += sessions_needed
            
            print(f"Course {course['code']}: {study_time_str} = {minutes} min = {sessions_needed} sessions")
        
        # Generate study time slots based on day type
        if not is_weekend:
            base_study_times = ["9:00 AM", "10:00 AM", "2:00 PM", "3:00 PM", "4:00 PM", "7:00 PM", "8:00 PM", "9:00 PM"]
        else:
            base_study_times = ["9:00 AM", "10:00 AM", "11:00 AM", "2:00 PM", "3:00 PM", "4:00 PM", "7:00 PM", "8:00 PM", "9:00 PM"]
        
        # Add more slots if needed for high study requirements
        extra_times = ["1:00 PM", "8:30 PM", "9:30 PM", "6:30 PM", "10:30 AM", "11:30 AM"]
        if total_daily_study_needed > len(base_study_times):
            base_study_times.extend(extra_times[:total_daily_study_needed - len(base_study_times)])
        
        # Schedule study sessions according to requirements
        study_slots_assigned = {}
        for course_code in course_study_requirements:
            study_slots_assigned[course_code] = 0
        
        # Get available study time slots
        available_times = []
        for study_time in base_study_times:
            if study_time in daily_schedule and daily_schedule[study_time]["type"] == "free":
                available_times.append(study_time)
        
        # Distribute study sessions fairly using round-robin approach
        max_sessions_per_course = max(course_study_requirements.values()) if course_study_requirements else 0
        
        for session_round in range(max_sessions_per_course):
            for course_code in course_study_requirements:
                if (study_slots_assigned[course_code] < course_study_requirements[course_code] and 
                    len(available_times) > 0):
                    
                    # Find next available time slot
                    for study_time in available_times[:]:  # Create a copy to modify
                        if daily_schedule[study_time]["type"] == "free":
                            daily_schedule[study_time] = {
                                "activity": f"{course_code} - Study Time",
                                "type": "study",
                                "date": current_day_date,
                                "course_code": course_code
                            }
                            study_slots_assigned[course_code] += 1
                            available_times.remove(study_time)
                            print(f"Assigned {course_code} study session {study_slots_assigned[course_code]} to {study_time}")
                            break
        
        # STEP 6: Add breaks (only in free slots)
        break_times = ["10:30 AM", "3:30 PM"]
        for break_time in break_times:
            if break_time in daily_schedule and daily_schedule[break_time]["type"] == "free":
                daily_schedule[break_time] = {"activity": "Break/Walk", "type": "break", "date": current_day_date}
        
        weekly_schedule[day] = daily_schedule
    
    return weekly_schedule

def create_schedule_dataframe(weekly_schedule):
    """Create a pandas DataFrame for the schedule table with dates"""
    time_slots = generate_time_slots()
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    # Get dates for the days (assuming they're stored in the schedule)
    day_headers = []
    for day in days:
        if day in weekly_schedule and weekly_schedule[day]:
            # Get date from any slot in that day
            first_slot = list(weekly_schedule[day].keys())[0]
            day_date = weekly_schedule[day][first_slot].get('date')
            if day_date:
                date_str = day_date.strftime('%m/%d')
                day_headers.append(f"{day}\n{date_str}")
            else:
                day_headers.append(day)
        else:
            day_headers.append(day)
    
    # Create the dataframe
    data = []
    for time_slot in time_slots:
        row = [time_slot]
        for day in days:
            slot_data = weekly_schedule.get(day, {}).get(time_slot, {"activity": "Free Time", "type": "free"})
            activity = slot_data["activity"]
            row.append(activity)
        data.append(row)
    
    df = pd.DataFrame(data, columns=["Time"] + day_headers)
    return df

def style_schedule_dataframe(df, weekly_schedule):
    """Apply light teal background with dark text for better readability"""
    
    def color_cell(val):
        # Handle blank cells
        if val == "":
            return "background-color: transparent; color: transparent;"
        
        # Default light teal background with dark text for all cells
        return "background-color: #b2f5ea; color: #2d3748; font-weight: 600; padding: 8px; border: 1px solid #81e6d9;"
    
    return df.style.applymap(color_cell, subset=df.columns[1:])

def generate_pdf_schedule(schedule_data, user_data):
    """Generate a PDF schedule in landscape with colors and better formatting"""
    buffer = BytesIO()
    # Use landscape orientation
    doc = SimpleDocTemplate(buffer, pagesize=(11*inch, 8.5*inch), rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    
    # Create custom styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=8,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#00d2d3')
    )
    
    wellness_style = ParagraphStyle(
        'Wellness',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=4,
        textColor=colors.HexColor('#2f3542'),
        leftIndent=15
    )
    
    # Build the story
    story = []
    
    # Title
    story.append(Paragraph("🎯 FocusFlow Weekly Template", title_style))
    story.append(Spacer(1, 8))
    
    # Important reminders section
    story.append(Paragraph("📋 Important Reminders", ParagraphStyle('RemindersHeader', parent=styles['Heading3'], fontSize=12, spaceAfter=6)))
    story.append(Paragraph("• <b>Sleep is crucial:</b> This schedule prioritizes 7-9 hours of sleep for optimal learning and memory consolidation.", wellness_style))
    story.append(Paragraph("• <b>Sample schedule:</b> This is one possible arrangement. Modify times and activities based on your needs and preferences.", wellness_style))
    story.append(Paragraph("• <b>Flexibility matters:</b> Use 'buffer time' between classes for walking, transitions, or short breaks.", wellness_style))
    story.append(Paragraph("• <b>Balance is key:</b> Study time is balanced with meals, exercise, and relaxation for long-term sustainability.", wellness_style))
    story.append(Spacer(1, 12))
    
    # Create schedule table with proper headers
    df = create_schedule_dataframe(schedule_data)
    
    # Use more time slots since we have more space in landscape
    key_times = ["6:00 AM", "7:00 AM", "8:00 AM", "9:00 AM", "10:00 AM", "11:00 AM", "12:00 PM", 
                "1:00 PM", "2:00 PM", "3:00 PM", "4:00 PM", "5:00 PM", "6:00 PM",
                "7:00 PM", "8:00 PM", "9:00 PM", "10:00 PM", "11:00 PM"]
    
    table_data = []
    # Add header row with proper day names
    header_row = ["Time"]
    for col in df.columns[1:]:  # Skip the Time column
        if '\n' in col:
            day_name = col.split('\n')[0]  # Get just the day name, not the date
        else:
            day_name = col
        header_row.append(day_name)
    table_data.append(header_row)
    
    # Add data rows with text wrapping
    for _, row in df.iterrows():
        if row['Time'] in key_times:
            data_row = [row['Time']]
            for col in df.columns[1:]:
                cell_content = str(row[col])
                # Truncate long text and add line breaks for better fitting
                if len(cell_content) > 20:
                    # Break long text into multiple lines
                    words = cell_content.split(' ')
                    lines = []
                    current_line = ""
                    for word in words:
                        if len(current_line + " " + word) <= 15:
                            current_line += (" " + word) if current_line else word
                        else:
                            if current_line:
                                lines.append(current_line)
                            current_line = word
                    if current_line:
                        lines.append(current_line)
                    cell_content = "<br/>".join(lines[:2])  # Max 2 lines
                    if len(lines) > 2:
                        cell_content += "..."
                
                # Wrap in Paragraph for proper text handling
                try:
                    cell_paragraph = Paragraph(cell_content, ParagraphStyle('CellText', fontSize=6, leading=8))
                    data_row.append(cell_paragraph)
                except:
                    # Fallback to plain text if Paragraph fails
                    data_row.append(str(cell_content))
            table_data.append(data_row)
    
    # Create table with better column widths for landscape and text wrapping
    table = Table(table_data, colWidths=[0.8*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch])
    
    # Apply colors to the table
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00d2d3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]
    
    # Apply cell-specific colors based on content
    for row_idx, row_data in enumerate(table_data[1:], 1):  # Skip header row
        for col_idx, cell_content in enumerate(row_data[1:], 1):  # Skip time column
            if cell_content:
                # Determine color based on activity type
                if 'Sleep' in cell_content and 'Go to' not in cell_content:
                    bg_color = colors.HexColor('#2f3542')
                elif 'Go to Sleep' in cell_content:
                    bg_color = colors.HexColor('#5a6c7d')
                elif 'Lecture' in cell_content or 'Lab' in cell_content:
                    bg_color = colors.HexColor('#3742fa')
                elif 'Study Time' in cell_content:
                    # Different colors for different courses
                    if 'MICROA' in cell_content or 'MICRO' in cell_content:
                        bg_color = colors.HexColor('#8e44ad')  # Purple
                    elif 'A&PI' in cell_content or 'A&P' in cell_content:
                        bg_color = colors.HexColor('#9b59b6')  # Light Purple
                    else:
                        bg_color = colors.HexColor('#5f27cd')  # Default study purple
                elif 'Breakfast' in cell_content or 'Lunch' in cell_content or 'Dinner' in cell_content:
                    bg_color = colors.HexColor('#ff9f43')
                elif 'Break' in cell_content:
                    bg_color = colors.HexColor('#ff6b6b')
                elif 'Practice' in cell_content or 'Game' in cell_content or 'Workout' in cell_content:
                    bg_color = colors.HexColor('#e67e22')
                elif 'Free Time' in cell_content:
                    bg_color = colors.HexColor('#00d2d3')
                else:
                    bg_color = colors.white
                
                # Add background color and white text for visibility
                table_style.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), bg_color))
                if bg_color != colors.white:
                    table_style.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.white))
    
    table.setStyle(TableStyle(table_style))
    story.append(table)
    story.append(Spacer(1, 12))
    
    # Evidence-based study strategies section
    story.append(Paragraph("📚 Evidence-Based Study Strategies", ParagraphStyle('StudyHeader', parent=styles['Heading3'], fontSize=12, spaceAfter=6)))
    story.append(Paragraph("<b>Active Recall:</b> Test yourself regularly instead of just re-reading. Use flashcards, practice questions, or explain concepts out loud.", wellness_style))
    story.append(Paragraph("<b>Spaced Repetition:</b> Review material at increasing intervals (1 day, 3 days, 1 week, 2 weeks) for better retention.", wellness_style))
    story.append(Paragraph("<b>Pomodoro Technique:</b> Study for 25 minutes, then take a 5-minute break. After 4 cycles, take a longer 15-30 minute break.", wellness_style))
    story.append(Paragraph("<b>Interleaving:</b> Mix different topics or types of problems in one study session rather than studying one subject for hours.", wellness_style))
    story.append(Paragraph("<b>Elaboration:</b> Connect new information to what you already know. Ask 'why' and 'how' questions.", wellness_style))
    story.append(Paragraph("<b>Dual Coding:</b> Use both visual and verbal information - draw diagrams, create mind maps, use charts and graphs.", wellness_style))
    story.append(Paragraph("<b>Practice Testing:</b> Take practice exams under timed conditions to simulate real testing scenarios.", wellness_style))
    story.append(Spacer(1, 8))
    
    # Color legend
    story.append(Paragraph("🎨 Color Guide", ParagraphStyle('ColorHeader', parent=styles['Heading3'], fontSize=12, spaceAfter=6)))
    story.append(Paragraph("Classes: Blue | Study Sessions: Purple variations by subject | Meals: Orange | Activities: Brown | Sleep: Dark Gray | Go to Sleep: Light Gray | Breaks: Red | Free Time: Teal", wellness_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def main():
    # Simple title
    st.markdown('<h1 class="app-title">🎯 FocusFlow</h1>', unsafe_allow_html=True)
    
    st.markdown('<div class="main-container">', unsafe_allow_html=True)
    
    # Step-by-step flow
    if st.session_state.step == 1:
        show_excel_upload()
    elif st.session_state.step == 2:
        show_preferences_step()
    elif st.session_state.step == 3:
        show_schedule_step()
    
    st.markdown("</div>", unsafe_allow_html=True)

def show_excel_upload():
    """Step 1: Excel file upload and processing"""
    st.markdown("""
    <div style="background: rgba(255, 255, 255, 0.08); border-radius: 12px; padding: 1rem; margin: 1rem 0; text-align: center;">
        <h2 style="color: #6c5ce7; margin-bottom: 0.5rem;">📚 Step 1: Course Information</h2>
        <p style="color: rgba(255, 255, 255, 0.8); margin-bottom: 0;">Upload Excel file, import configuration, or add courses manually</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Main course upload section first
    if st.session_state.get('editing_course') is None and not st.session_state.get('file_processed', False):
        st.markdown("### 📄 Upload Course Template")
        
        # Instructions
        st.info("""
        **Template Format Requirements:**
        - Excel file with separate tabs for each course (Course 1, Course 2, etc.)
        - Each tab should contain course information, class schedules, and assignment details
        - Use the format: Course title, lecture schedule, lab schedule, assignments, etc.
        - **⏰ IMPORTANT: All times must include AM or PM (e.g., "9:00 AM", "2:30 PM")**
        """)
        
        st.warning("⏰ **Time Format Note:** Make sure all class times in your Excel file include AM or PM (e.g., '10:30 AM - 11:30 AM')")
        
        uploaded_file = st.file_uploader(
            "Choose Excel file",
            type=['xlsx', 'xls'],
            help="Upload your course template Excel file"
        )
        
        if uploaded_file is not None:
            with st.spinner("🧠 Processing your course template..."):
                courses = parse_excel_course_file(uploaded_file)
                
                if courses:
                    st.session_state.courses = courses
                    
                    # Create assignments list
                    all_assignments = []
                    for course in courses:
                        all_assignments.extend(course.get('assignments', []))
                    st.session_state.assignments = all_assignments
                    
                    st.success(f"✅ Successfully loaded {len(courses)} courses with {len(all_assignments)} assignments!")
                    
                    # Mark file as processed to prevent reprocessing
                    st.session_state.file_processed = True
                    st.rerun()
                else:
                    st.error("❌ Could not read course data from the file. Please check the format.")
        
        # Alternative options section
        st.markdown("---")
        st.markdown("### 📥 Alternative Options")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Download template button
            template_buffer = create_excel_template()
            st.download_button(
                label="📄 Download Excel Template",
                data=template_buffer.getvalue(),
                file_name="FocusFlow_Course_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download a sample Excel template to fill out with your course information"
            )
        
        with col2:
            # Import configuration
            uploaded_config = st.file_uploader(
                "Import Saved Configuration",
                type=['json'],
                help="Upload a previously saved FocusFlow configuration file",
                key="config_upload"
            )
            
            if uploaded_config is not None:
                try:
                    config_data = uploaded_config.read().decode('utf-8')
                    if load_config_from_json(config_data):
                        st.success("✅ Configuration loaded successfully!")
                        # Skip to step 2 if user data is also loaded
                        if st.session_state.get('user_data'):
                            if st.button("🚀 Go to Schedule Generation"):
                                st.session_state.step = 3
                                # Generate schedule with loaded data
                                schedule = generate_weekly_schedule(
                                    st.session_state.courses,
                                    st.session_state.intramurals,
                                    st.session_state.user_data
                                )
                                st.session_state.final_schedule = schedule
                                st.rerun()
                        else:
                            if st.button("➡️ Continue to Preferences"):
                                st.session_state.step = 2
                                st.rerun()
                    else:
                        st.error("❌ Could not load configuration file. Please check the format.")
                except Exception as e:
                    st.error(f"❌ Error loading configuration: {str(e)}")
    
    # Show manual course addition option when no file has been processed
    if st.session_state.get('editing_course') is None and not st.session_state.get('file_processed', False):
        # Option to add course manually
        with st.expander("➕ Add Course Manually"):
            st.markdown("**Add a course if you don't have an Excel file:**")
            
            col1, col2 = st.columns(2)
            with col1:
                manual_code = st.text_input("Course Code", key="manual_code", placeholder="e.g., BIO1205")
                manual_name = st.text_input("Course Name", key="manual_name", placeholder="e.g., Biology Lab")
            
            with col2:
                manual_study_time = st.text_input("Daily Study Time", key="manual_study", placeholder="e.g., 30 min")
            
            # Class schedule
            st.markdown("**Class Schedule:**")
            st.caption("⏰ Remember: Include AM or PM in all times (e.g., 9:00 AM)")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                manual_days = st.multiselect(
                    "Days",
                    ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
                    key="manual_days"
                )
            
            with col2:
                manual_start = st.text_input("Start Time", key="manual_start", placeholder="e.g., 9:00 AM")
                st.caption("Include AM/PM")
            
            with col3:
                manual_end = st.text_input("End Time", key="manual_end", placeholder="e.g., 10:30 AM")
                st.caption("Include AM/PM")
            
            with col4:
                manual_location = st.text_input("Location", key="manual_location", placeholder="e.g., Room 101")
            
            if st.button("Add Manual Course"):
                if manual_code and manual_name:
                    manual_course = {
                        'code': manual_code,
                        'name': manual_name,
                        'daily_study_time': manual_study_time or '30 min',
                        'class_schedule': []
                    }
                    
                    if manual_days and manual_start and manual_end:
                        manual_course['class_schedule'] = [{
                            'days': manual_days,
                            'start_time': manual_start,
                            'end_time': manual_end,
                            'type': 'Lecture',
                            'location': manual_location
                        }]
                    
                    st.session_state.courses.append(manual_course)
                    st.session_state.file_processed = True  # Mark as having course data
                    st.success(f"✅ Added {manual_code} manually!")
                    st.rerun()
                else:
                    st.error("Please fill in at least the course code and name.")
    
    # Show current courses if any
    if st.session_state.courses:
        st.markdown("### 📚 Your Courses")
        for i, course in enumerate(st.session_state.courses):
            with st.expander(f"📖 {course.get('code', 'Unknown')} - {course.get('name', 'Unknown Course')}", expanded=False):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Course Information:**")
                    st.write(f"**Code:** {course.get('code', 'Unknown')}")
                    st.write(f"**Name:** {course.get('name', 'Unknown Course')}")
                    st.write(f"**Daily Study Time:** {course.get('daily_study_time', 'Not specified')}")
                
                with col2:
                    st.markdown("**Class Schedule:**")
                    if 'class_schedule' in course and course['class_schedule']:
                        for class_time in course['class_schedule']:
                            days_str = ", ".join(class_time['days'])
                            location = f" - {class_time['location']}" if class_time.get('location') else ""
                            st.write(f"• **{class_time['type']}:** {days_str}, {class_time['start_time']} - {class_time['end_time']}{location}")
                    else:
                        st.write("No class schedule found")
                
                # Show assignments if any
                if course.get('assignments'):
                    st.markdown("**Assignments:**")
                    for assignment in course['assignments']:
                        st.write(f"• {assignment['title']} - {assignment['date']} ({assignment['type']})")
                
                # Option to edit this course
                if st.button(f"✏️ Edit {course.get('code', 'Course')}", key=f"edit_{i}"):
                    st.session_state.editing_course = i
                    st.rerun()
    
    # Show editing interface if a course is being edited
    if st.session_state.get('editing_course') is not None:
        course_index = st.session_state.editing_course
        course = st.session_state.courses[course_index]
        
        st.markdown("### ✏️ Edit Course")
        
        # Basic course info
        col1, col2 = st.columns(2)
        with col1:
            new_code = st.text_input("Course Code", value=course.get('code', ''), key="edit_code")
            new_name = st.text_input("Course Name", value=course.get('name', ''), key="edit_name")
        
        with col2:
            new_study_time = st.text_input("Daily Study Time", value=course.get('daily_study_time', '30 min'), key="edit_study")
        
        # Class schedule editing
        st.markdown("**Class Schedule:**")
        if 'class_schedule' in course:
            for j, class_time in enumerate(course['class_schedule']):
                st.markdown(f"**{class_time['type']} {j+1}:**")
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    new_days = st.multiselect(
                        "Days",
                        ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
                        default=class_time.get('days', []),
                        key=f"edit_days_{j}"
                    )
                
                with col2:
                    new_start = st.text_input("Start Time", value=class_time.get('start_time', ''), key=f"edit_start_{j}")
                    st.caption("Include AM/PM")
                
                with col3:
                    new_end = st.text_input("End Time", value=class_time.get('end_time', ''), key=f"edit_end_{j}")
                    st.caption("Include AM/PM")
                
                with col4:
                    new_location = st.text_input("Location", value=class_time.get('location', ''), key=f"edit_location_{j}")
                
                # Update the class schedule
                course['class_schedule'][j] = {
                    'days': new_days,
                    'start_time': new_start,
                    'end_time': new_end,
                    'type': class_time['type'],
                    'location': new_location
                }
        
        # Save/Cancel buttons
        col1, col2 = st.columns(2)
        with col1:
            if st.button("💾 Save Changes", key="save_course"):
                course['code'] = new_code
                course['name'] = new_name
                course['daily_study_time'] = new_study_time
                st.session_state.courses[course_index] = course
                st.session_state.editing_course = None
                st.success("✅ Course updated!")
                st.rerun()
        
        with col2:
            if st.button("❌ Cancel", key="cancel_edit"):
                st.session_state.editing_course = None
                st.rerun()
    
    # File upload section (only show if not editing and haven't processed a file yet)
    if st.session_state.get('editing_course') is None and not st.session_state.get('file_processed', False):
        # This section is now moved to the top of the function
        pass
    
    # Show manual course addition option when no file has been processed
    if st.session_state.get('editing_course') is None and not st.session_state.get('file_processed', False):
        # This section is now moved to the top of the function  
        pass
    
    # Show action buttons when courses are loaded
    if st.session_state.courses and st.session_state.get('editing_course') is None:
        st.markdown("### 🎯 Ready to Continue?")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📝 Add Another Course"):
                # Allow them to add more courses manually
                with st.form("add_course_form"):
                    st.markdown("**Add Another Course:**")
                    
                    form_col1, form_col2 = st.columns(2)
                    with form_col1:
                        new_code = st.text_input("Course Code", placeholder="e.g., CHEM101")
                        new_name = st.text_input("Course Name", placeholder="e.g., General Chemistry")
                    
                    with form_col2:
                        new_study_time = st.text_input("Daily Study Time", placeholder="e.g., 45 min")
                    
                    # Class schedule
                    st.markdown("**Class Schedule:**")
                    st.caption("⏰ Remember: Include AM or PM in all times (e.g., 11:00 AM)")
                    form_col1, form_col2, form_col3, form_col4 = st.columns(4)
                    
                    with form_col1:
                        new_days = st.multiselect(
                            "Days",
                            ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                        )
                    
                    with form_col2:
                        new_start = st.text_input("Start Time", placeholder="e.g., 11:00 AM")
                        st.caption("Include AM/PM")
                    
                    with form_col3:
                        new_end = st.text_input("End Time", placeholder="e.g., 12:30 PM")
                        st.caption("Include AM/PM")
                    
                    with form_col4:
                        new_location = st.text_input("Location", placeholder="e.g., Chem Lab")
                    
                    if st.form_submit_button("Add Course"):
                        if new_code and new_name:
                            new_course = {
                                'code': new_code,
                                'name': new_name,
                                'daily_study_time': new_study_time or '30 min',
                                'class_schedule': []
                            }
                            
                            if new_days and new_start and new_end:
                                new_course['class_schedule'] = [{
                                    'days': new_days,
                                    'start_time': new_start,
                                    'end_time': new_end,
                                    'type': 'Lecture',
                                    'location': new_location
                                }]
                            
                            st.session_state.courses.append(new_course)
                            st.success(f"✅ Added {new_code}!")
                            st.rerun()
        
        with col2:
            if st.button("🚀 Continue to Preferences", type="primary", key="continue_main"):
                st.session_state.step = 2
                st.rerun()
    
    # Progress and navigation
    st.markdown("""
    <div class="progress-bar">
        <div class="progress-fill" style="width: 33%"></div>
    </div>
    <p class="progress-text">Step 1 of 3 - Upload your course template to continue</p>
    """, unsafe_allow_html=True)
    
    # Navigation buttons at bottom
    col1, col2 = st.columns(2)
    
    with col1:
        if st.session_state.courses:
            if st.button("🔄 Upload Different File"):
                # Reset everything to start over
                st.session_state.courses = []
                st.session_state.assignments = []
                st.session_state.file_processed = False
                st.session_state.editing_course = None
                st.rerun()
    
    with col2:
        # This button is now handled above in the action section
        if not st.session_state.courses:
            st.info("👆 Upload your course template or add courses manually to continue")


def show_preferences_step():
    """Step 2: Preferences setup"""
    st.markdown("""
    <div style="background: rgba(255, 255, 255, 0.08); border-radius: 12px; padding: 1rem; margin: 1rem 0; text-align: center;">
        <h2 style="color: #00d2d3; margin-bottom: 0.5rem;">⚙️ Step 2: Your Preferences</h2>
        <p style="color: rgba(255, 255, 255, 0.8); margin-bottom: 0;">Set your schedule preferences</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### ⏰ Time Preferences")
        wake_time = st.slider("Wake up time", 6, 11, 8, format="%d:00 AM")
        
        # Bedtime with AM/PM display
        bedtime_hour = st.slider("Bedtime hour", 9, 2, 11)
        if bedtime_hour >= 9:
            bedtime_display = f"{bedtime_hour}:00 PM"
            bedtime_actual = bedtime_hour  # 21, 22, 23 for 9PM, 10PM, 11PM
        else:
            bedtime_display = f"{bedtime_hour}:00 AM"
            bedtime_actual = bedtime_hour + 24  # 25, 26 for 1AM, 2AM
        
        st.write(f"**Bedtime: {bedtime_display}**")
        
        st.markdown("### 📅 Schedule Timeline")
        schedule_type = st.selectbox(
            "What type of schedule do you want?",
            ["📅 Current Week (this week)", "📊 Template Week (reusable)", "🗓️ Custom Start Date"]
        )
        
        start_date = datetime.now().date()
        if schedule_type == "🗓️ Custom Start Date":
            start_date = st.date_input(
                "Choose start date for your schedule:",
                value=datetime.now().date(),
                help="Pick the Monday you want your schedule to start"
            )
            # Adjust to Monday if not already
            days_since_monday = start_date.weekday()
            start_date = start_date - timedelta(days=days_since_monday)
            st.info(f"Schedule will start on Monday: {start_date.strftime('%B %d, %Y')}")
        elif schedule_type == "📊 Template Week (reusable)":
            st.info("Creates a generic weekly template without specific dates")
        
        st.markdown("### 📚 Study Preferences")
        study_intensity = st.selectbox(
            "Study intensity",
            ["🌿 Light (2-3 sessions/day)", "⚖️ Moderate (3-4 sessions/day)", "🔥 Intensive (4-5 sessions/day)"]
        )
    
    with col2:
        st.markdown("### 🏃 Intramural Activities")
        include_intramurals = st.checkbox("Include intramural/exercise time", value=False)
        
        if include_intramurals:
            st.markdown("**Add Your Activities:**")
            activity_name = st.text_input("Activity Name (e.g., Soccer, Basketball)", key="activity_name")
            activity_type = st.selectbox("Activity Type", ["Practice", "Game", "Workout", "Club Meeting", "Other"], key="activity_type")
            
            # Activity scheduling
            is_scheduled = st.checkbox("Has specific schedule?", value=False, key="is_scheduled")
            
            if is_scheduled:
                selected_days = st.multiselect(
                    "Select days",
                    ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
                    key="selected_days"
                )
                activity_start_time = st.time_input("Start Time", value=time(17, 0), key="activity_start_time")
                activity_duration = st.slider("Duration (minutes)", 30, 180, 90, key="activity_duration")
                
                st.write(f"**Schedule: {', '.join(selected_days)} at {activity_start_time.strftime('%I:%M %p')} for {activity_duration} minutes**")
            
            if st.button("Add Activity", key="add_activity") and activity_name:
                intramural = {
                    'name': activity_name,
                    'type': activity_type,
                    'scheduled': is_scheduled,
                    'days': selected_days if is_scheduled else [],
                    'start_time': activity_start_time.strftime('%I:%M %p') if is_scheduled else None,
                    'duration': activity_duration if is_scheduled else 60
                }
                st.session_state.intramurals.append(intramural)
                st.success(f"✅ Added {activity_name}!")
                
                # Debug: Show what was added
                st.write(f"**Debug Info:** Added activity with scheduled={is_scheduled}, days={selected_days if is_scheduled else []}")
                
                st.rerun()
        
        # Show added intramurals
        if st.session_state.intramurals:
            st.markdown("""
            <div class="intramural-card">
                <h4>🏃 Your Activities</h4>
            </div>
            """, unsafe_allow_html=True)
            
            for i, intramural in enumerate(st.session_state.intramurals):
                schedule_info = ""
                if intramural.get('scheduled'):
                    days = ", ".join(intramural.get('days', []))
                    start_time = intramural.get('start_time', 'TBD')
                    duration = intramural.get('duration', 60)
                    schedule_info = f" - {days} at {start_time} ({duration} min)"
                
                st.markdown(f"""
                <div class="activity-item">
                    <strong>{intramural['name']}</strong> ({intramural['type']}){schedule_info}
                    <br><small>Scheduled: {intramural.get('scheduled', False)} | Days: {intramural.get('days', [])}</small>
                </div>
                """, unsafe_allow_html=True)
                
                # Add delete button for each activity
                if st.button(f"🗑️ Delete {intramural['name']}", key=f"delete_activity_{i}"):
                    st.session_state.intramurals.pop(i)
                    st.rerun()
    
    # Show upcoming assignments if any
    if st.session_state.get('assignments'):
        st.markdown("### 📋 Upcoming Assignments")
        
        # Parse and sort assignments by date
        upcoming_assignments = []
        start_date = datetime.now().date()
        for assignment in st.session_state.assignments:
            try:
                # Try to parse the date
                date_str = assignment.get('date', '')
                if date_str and date_str.lower() != 'n/a':
                    # Handle various date formats
                    possible_formats = ['%m/%d/%Y', '%m-%d-%Y', '%Y-%m-%d', '%B %d, %Y', '%b %d, %Y']
                    parsed_date = None
                    
                    for fmt in possible_formats:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt).date()
                            break
                        except:
                            continue
                    
                    if parsed_date and parsed_date >= start_date:
                        weeks_from_start = (parsed_date - start_date).days // 7
                        if weeks_from_start < 4:  # Only show assignments in the 4-week period
                            assignment['parsed_date'] = parsed_date
                            assignment['week_number'] = weeks_from_start + 1
                            upcoming_assignments.append(assignment)
            except:
                continue
        
        if upcoming_assignments:
            upcoming_assignments.sort(key=lambda x: x['parsed_date'])
            
            for assignment in upcoming_assignments[:10]:  # Show first 10
                date_str = assignment['parsed_date'].strftime('%B %d, %Y')
                week_str = f"Week {assignment['week_number']}"
                priority_emoji = "🔴" if assignment.get('priority') == 'high' else "🟡" if assignment.get('priority') == 'medium' else "🟢"
                
                st.markdown(f"""
                <div style="background: rgba(108, 92, 231, 0.1); padding: 0.5rem; margin: 0.25rem 0; border-radius: 8px; border-left: 3px solid #6c5ce7;">
                    {priority_emoji} <strong>{assignment['course']}</strong> - {assignment['title']}<br>
                    <small>📅 {date_str} ({week_str})</small>
                </div>
                """, unsafe_allow_html=True)
    
    # Wellness information
    st.markdown("""
    <div class="wellness-info">
        <h4>🌟 Wellness Focus (Automatically Included)</h4>
        <p><strong>Sleep:</strong> 7-9 hours prioritized for memory consolidation and focus</p>
        <p><strong>Meals:</strong> Regular eating times maintain energy and brain function</p>
        <p><strong>Free Time:</strong> Social time and relaxation prevent burnout</p>
        <p><strong>Balance:</strong> Sustainable schedule for long-term success</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Store preferences (using bedtime_actual for calculations)
    preferences = {
        'wake_time': wake_time,
        'bedtime': bedtime_actual,
        'study_intensity': study_intensity,
        'include_intramurals': include_intramurals,
        'schedule_type': schedule_type,
        'start_date': start_date
    }
    
    st.session_state.user_data = preferences
    
    # Progress
    st.markdown("""
    <div class="progress-bar">
        <div class="progress-fill" style="width: 66%"></div>
    </div>
    <p class="progress-text">Step 2 of 3</p>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("← Back to Upload"):
            st.session_state.step = 1
            st.rerun()
    
    with col2:
        if st.button("🎨 Generate Schedule", type="primary"):
            with st.spinner("🎨 Creating your personalized schedule..."):
                schedule = generate_weekly_schedule(
                    st.session_state.courses,
                    st.session_state.intramurals,
                    preferences
                )
                st.session_state.final_schedule = schedule
                st.session_state.step = 3
                st.rerun()

def show_schedule_step():
    """Step 3: Schedule display with color-coded table"""
    st.markdown("""
    <div style="background: rgba(255, 255, 255, 0.08); border-radius: 12px; padding: 1rem; margin: 1rem 0; text-align: center;">
        <h2 style="color: #6c5ce7; margin-bottom: 0.5rem;">📅 Your Weekly Schedule Template</h2>
        <p style="color: rgba(255, 255, 255, 0.8); margin-bottom: 0;">A balanced weekly template you can repeat</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Compact stats in one row
    courses_count = len(st.session_state.courses)
    intramurals_count = len(st.session_state.intramurals)
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Courses", courses_count)
    with col2:
        st.metric("Activities", intramurals_count)
    with col3:
        st.metric("Weekdays Study", "4 sessions")
    with col4:
        st.metric("Weekend Study", "6 sessions")
    
    # Compact legend updated for new color scheme
    st.markdown("""
    <div style="text-align: center; margin: 0.5rem 0; font-size: 0.8rem;">
        <span style="background: #b2f5ea; color: #2d3748; padding: 4px 8px; border-radius: 4px; margin: 2px;">All Activities</span> - 
        Easy to read light teal background for all schedule items
    </div>
    """, unsafe_allow_html=True)
    
    # Display the weekly schedule
    if st.session_state.final_schedule:
        df = create_schedule_dataframe(st.session_state.final_schedule)
        styled_df = style_schedule_dataframe(df, st.session_state.final_schedule)
        
        # Debug section - show what activities were processed
        if st.session_state.intramurals:
            with st.expander("🔍 Debug: Activity Processing"):
                st.write("**Intramural activities in session state:**")
                for i, activity in enumerate(st.session_state.intramurals):
                    st.write(f"{i+1}. {activity}")
                
                # Check if activities appear in the schedule
                st.write("**Activities found in schedule:**")
                activity_found = False
                for day, schedule in st.session_state.final_schedule.items():
                    for time_slot, slot_data in schedule.items():
                        if slot_data["type"] == "activity":
                            st.write(f"- {day} {time_slot}: {slot_data['activity']}")
                            activity_found = True
                
                if not activity_found:
                    st.error("❌ No activities found in the generated schedule!")
                    st.write("This suggests an issue with the scheduling algorithm.")
        
        # Display the schedule table
        st.dataframe(
            styled_df,
            use_container_width=True,
            hide_index=True,
            height=800
        )
    
    # Export section
    st.markdown("### 🚀 Export Your Schedule")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.session_state.final_schedule and st.session_state.user_data:
            # Generate PDF
            pdf_buffer = generate_pdf_schedule(st.session_state.final_schedule, st.session_state.user_data)
            pdf_data = pdf_buffer.getvalue()
            
            st.download_button(
                label="📄 Download PDF Schedule",
                data=pdf_data,
                file_name=f"FocusFlow_Weekly_Template_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                help="Download your weekly schedule template with study strategies"
            )
    
    with col2:
        if st.session_state.courses:
            # Create save data
            save_data = {
                'courses': st.session_state.courses,
                'intramurals': st.session_state.intramurals,
                'user_data': st.session_state.user_data,
                'created_date': datetime.now().strftime('%Y-%m-%d'),
                'app_version': '1.0'
            }
            
            save_json = json.dumps(save_data, indent=2, default=str)
            
            st.download_button(
                label="💾 Save Configuration",
                data=save_json,
                file_name=f"FocusFlow_Config_{datetime.now().strftime('%Y%m%d')}.json",
                mime="application/json",
                help="Save your setup to reload later if you need to make changes"
            )
    
    # Navigation
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🔄 Modify Schedule"):
            st.session_state.step = 2
            st.rerun()
    
    with col2:
        if st.button("🏠 Start Over"):
            # Reset all session state
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    
    # Progress complete
    st.markdown("""
    <div class="progress-bar">
        <div class="progress-fill" style="width: 100%"></div>
    </div>
    <p class="progress-text">🎉 Weekly Template Complete!</p>
    """, unsafe_allow_html=True)
    
    st.success(f"""
    🎉 **Your Weekly Schedule Template is Ready!**
    
    ✅ **{courses_count} courses** with actual class times
    ✅ **{intramurals_count} activities** scheduled properly
    ✅ **Balanced study time** with course variety
    ✅ **30-minute increments** for precise planning
    ✅ **Sleep hygiene** with "Go to Sleep" reminders
    ✅ **Reusable template** for any week
    
    📚 **Use this as your weekly template and adjust as needed for each week!**
    """)

if __name__ == "__main__":
    main()
